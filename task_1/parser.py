import json
import openpyxl as xl
from openpyxl.reader import excel

def parser(json_file_name):
    with open(json_file_name, 'r+', encoding='utf-8') as json_file:
        json_dict = json.loads(json_file.read())
        return json_dict


def excel_load(json_dict, workbook):
    last_sheet_name = workbook.sheetnames[-1]
    ws = workbook[last_sheet_name]
    headers = json_dict["headers"]
    values = json_dict["values"]
    for col in range(1, len(headers)+1):
        header = headers[col-1]
        _ = ws.cell(column=col, row=1, value=header["properties"]["QuickInfo"])
        row = 2
        for i in range(len(values)):
            if values[i]["properties"]["X"] == header["properties"]["X"]:
                _ = ws.cell(column=col, row=row, value=values[i]["properties"]["Text"])
                row += 1
    


wb = xl.Workbook()
files_names = ["test1.json", "test2.json", "test3.json"]

for i in range(len(files_names)):
    json_dict = parser(files_names[i])
    excel_load(json_dict, wb)
    if i != len(files_names)-1:
        wb.create_sheet()

wb.save('test_xl.xslx')